import os
import sys
import requests
from bs4 import BeautifulSoup
import re
from collections import defaultdict
from datetime import datetime
import openpyxl
import tkinter as tk
from tkinter import messagebox  


if getattr(sys, 'frozen', False):
    # If the application is running as a bundled executable
    script_dir = os.path.dirname(sys.executable)
else:
    # If the application is running as a Python script
    script_dir = os.path.dirname(os.path.abspath(__file__))

root = tk.Tk()
root.withdraw()  
root.attributes('-topmost', True)  # Ensure the popup is on top

data_file_path = os.path.join(script_dir, 'data.txt')
result_folder = os.path.join(script_dir, 'result')
os.makedirs(result_folder, exist_ok=True)

try:
    with open(data_file_path, 'r') as file:
        lines = file.readlines()
except FileNotFoundError:
    messagebox.showerror("File Not Found", f"Error: {data_file_path} does not exist. Please ensure the file is present.")
    sys.exit()  

if len(lines) < 2:
    messagebox.showerror("Invalid Data", "The data.txt file must contain at least 2 lines.")
    sys.exit()

start_date_values = lines[0].strip().split(', ')
end_date_values = lines[1].strip().split(', ')
user_urls = [line.strip() for line in lines[2:] if line.strip()]

if len(start_date_values) != 3 or len(end_date_values) != 3:
    print("data type must be 'YYYY, MM, DD' provided in data.txt")
    sys.exit()

try:
    start_date = datetime(int(start_date_values[0]), int(start_date_values[1]), int(start_date_values[2]))
    end_date = datetime(int(end_date_values[0]), int(end_date_values[1]), int(end_date_values[2]))
except ValueError:
    print("Date Type is wrong. It requires 'YYYY, MM, DD'.")
    sys.exit()

if not user_urls:
    print("URL list is empty. Fill the data.txt with URLs.")
    sys.exit()

player_names = [
    'Alisa', 'Asuka', 'Azucena', 'Bryan', 'Claudio', 'Devil Jin', 'Dragunov', 'Feng',
    'Hwoarang', 'Jack-8', 'Jin', 'Jun', 'Kazuya', 'King', 'Kuma', 'Lars', 'Lee',
    'Leo', 'Leroy', 'Lili', 'Xiaoyu', 'Law', 'Nina', 'Panda', 'Paul', 'Raven',
    'Reina', 'Shaheen', 'Steve', 'Victor', 'Yoshimitsu', 'Zafina', 'Eddy', 'Lidia', "Heihachi"
]

def clean_filename(filename):
    """ Clean the filename by replacing invalid characters with underscores. """
    return re.sub(r'[<>:"/\\|?*]', '_', filename)

def clean_text(text):
    text = text.replace('\n', ' ')
    return re.sub(r'\s+', ' ', text).strip()

def convert_datetime_format(datetime_str):
    try:
        datetime_obj = datetime.strptime(datetime_str, '%d %b %Y %H:%M')
        return datetime_obj.strftime('%Y-%m-%d %H:%M')
    except ValueError:
        return datetime_str

def is_within_date_range(date_str):
    try:
        date_obj = datetime.strptime(date_str, '%d %b %Y %H:%M')
        return start_date <= date_obj <= end_date
    except ValueError:
        return False

for user in user_urls:
    response = requests.get(user)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    sheet_papyrus_divs = soup.find_all('div', class_='sheet papyrus') 
    
    title_tag = soup.find('title')
    if title_tag:
        full_title = title_tag.get_text()
        title_parts = full_title.split('•')
        title = title_parts[0].strip() if title_parts else full_title.strip()
    else:
        title = 'No title found'

    categorized_data = defaultdict(list)
    statistics = defaultdict(lambda: defaultdict(lambda: {'WIN': 0, 'LOSE': 0, 'DRAW': 0, 'total': 0}))
    total_statistics = defaultdict(lambda: {'WIN': 0, 'LOSE': 0, 'DRAW': 0, 'total': 0})

    for div in sheet_papyrus_divs:
        tables = div.find_all('table')  
        for table in tables:
            rows = table.find_all('tr')  
            for row in rows:
                cells = row.find_all('td')
                cell_values = [clean_text(cell.get_text()) for cell in cells]

                if len(cell_values) > 1:
                    datetime_str = cell_values[0]
                    if is_within_date_range(datetime_str):
                        converted_datetime_str = convert_datetime_format(datetime_str)
                        cell_values[0] = converted_datetime_str

                        key = cell_values[1]
                        categorized_data[key].append(cell_values)

                        opponent = cell_values[5]
                        result = cell_values[2].split()[1]
                        if opponent in player_names:
                            statistics[key][opponent][result] += 1
                            statistics[key][opponent]['total'] += 1

                            total_statistics[opponent][result] += 1
                            total_statistics[opponent]['total'] += 1

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        clean_title = clean_filename(title)  # Clean title for the filename
        filename = f'{clean_title}_{start_date.strftime("%Y_%m_%d")}_{end_date.strftime("%Y_%m_%d")}_{timestamp}.xlsx'
        filepath = os.path.join(result_folder, filename)

        workbook = openpyxl.Workbook()
        total_sheet = workbook.create_sheet(title='Total')

        for player in player_names:
            if player in statistics:
                sheet = workbook.create_sheet(title=player)
                headers = ['Opponent', 'Wins', 'Losses', 'Draws', 'Win Rate']
                sheet.append(headers)

                total_wins = total_losses = total_draws = total_matches = 0

                for opponent in player_names:
                    if opponent in statistics[player]:
                        stats = statistics[player][opponent]
                        wins = stats['WIN']
                        losses = stats['LOSE']
                        draws = stats['DRAW']
                        total = stats['total']
                        win_rate = (wins / total) * 100 if total > 0 else 0
                        row = [opponent, wins, losses, draws, f'{win_rate:.2f}%']
                        sheet.append(row)
                        total_wins += wins
                        total_losses += losses
                        total_draws += draws
                        total_matches += total

                sheet.append(['Total', total_wins, total_losses, total_draws, f'{(total_wins / total_matches) * 100 if total_matches > 0 else 0:.2f}%'])

        total_headers = ['Opponent', 'Wins', 'Losses', 'Draws', 'Win Rate']
        total_sheet.append(total_headers)

        grand_total_wins = grand_total_losses = grand_total_draws = grand_total_matches = 0

        for opponent in player_names:
            if opponent in total_statistics:
                stats = total_statistics[opponent]
                wins = stats['WIN']
                losses = stats['LOSE']
                draws = stats['DRAW']
                total = stats['total']
                win_rate = (wins / total) * 100 if total > 0 else 0
                total_sheet.append([opponent, wins, losses, draws, f'{win_rate:.2f}%'])
                grand_total_wins += wins
                grand_total_losses += losses
                grand_total_draws += draws  # Fixed here
                grand_total_matches += total

        total_sheet.append(['Total', grand_total_wins, grand_total_losses, grand_total_draws, f'{(grand_total_wins / grand_total_matches) * 100 if grand_total_matches > 0 else 0:.2f}%'])

        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        workbook.save(filepath)

messagebox.showinfo("Complete", "Data analysis has been completed successfully. Check Result Folder")
root.quit()
